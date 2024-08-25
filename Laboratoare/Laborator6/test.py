import openpyxl

# Define variable to load the dataframe
dataframe = openpyxl.load_workbook("test.xlsx")

# Define variable to read sheet
dataframe1 = dataframe.active

grades = []
target=8.5*4/5+9.43*1/5
# Iterate the loop to read the cell values
for row in range(0, dataframe1.max_row):
    i=0
    for col in dataframe1.iter_cols(3, 4):
        if type(col[row].value) is float:
            if i == 0:
                grades.append(0.2*col[row].value)
            else:
                grades[len(grades)-1] += 4/5*col[row].value
            i=i+1

sortedGrades = sorted(grades,reverse=True)
print(sortedGrades)
print(len([x for x in grades if x >= target]))